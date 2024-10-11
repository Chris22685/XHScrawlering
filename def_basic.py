import openpyxl
from DrissionPage.errors import ElementNotFoundError
import random
import time
import re
from datetime import datetime, timedelta


'''读取搜索文章结果函数'''
def get_info(search_page, contents):
    # 定位包含笔记信息的sections
    container = search_page.ele('.feeds-page')
    sections = container.eles('.note-item')
    for section in sections:
        # 定位文章链接
        try:
            note_link = section.ele('tag:a', timeout=0).link
        except ElementNotFoundError:
            print("文章链接元素未找到。跳过当前循环...")
            continue

        # 定位标题
        footer = section.ele('.footer', timeout=0.2)
        title = "空标题"
        try:
            title = footer.ele('.title', timeout=0.2).text
            if not title:  # 如果标题为空，跳过当前循环
                title = "空标题"
        except ElementNotFoundError:
            print("Title element not found. Skipping...")
            title = "空标题"
        # 定位作者
        author_wrapper = footer.ele('.author-wrapper')
        author = author_wrapper.ele('.author').text
        # 定位作者主页地址
        author_link = author_wrapper.ele('tag:a', timeout=0).link
        # 定位作者头像
        author_img = author_wrapper.ele('tag:img', timeout=0).link
        # 定位点赞
        like = footer.ele('.like-wrapper like-active').text
        # 将获取到的信息添加到 contents 列表
        contents.append([title, author, note_link, author_link, author_img, like])

'''下滑页面函数'''
def page_scroll_down(search_page):
    print("********下滑页面********")
    # 生成一个随机时间
    random_time = random.uniform(0.5, 1.5)
    # 暂停
    time.sleep(random_time)
    # time.sleep(1)
    # page.scroll.down(5000)
    search_page.scroll.to_bottom()

'''整理Excel表格函数'''
def auto_resize_column(excel_path):
    """自适应列宽度"""
    wb = openpyxl.load_workbook(excel_path)
    worksheet = wb.active
    # 循环遍历工作表中的1-2列
    for col in worksheet.iter_cols(min_col=1, max_col=2):
        max_length = 0
        # 列名称
        column = col[0].column_letter
        # 循环遍历列中的所有单元格
        for cell in col:
            try:
                # 如果当前单元格的值长度大于max_length，则更新 max_length 的值
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # 计算调整后的列宽度
        adjusted_width = (max_length + 2) * 2
        # 使用 worksheet.column_dimensions 属性设置列宽度
        worksheet.column_dimensions[column].width = adjusted_width

        # 循环遍历工作表中的3-5列
        for col in worksheet.iter_cols(min_col=3, max_col=11):
            max_length = 0
            column = col[0].column_letter  # Get the column name

            # 使用 worksheet.column_dimensions 属性设置列宽度
            worksheet.column_dimensions[column].width = 15

    wb.save(excel_path)

"""将'XXX万'格式的数量转换为整数"""
def convert_count_to_int(count_str):
    """
    将包含'万'的数字字符串转换为整数
    
    参数:
        count_str: 字符串，例如 '1.1万', '2066'
    
    返回:
        整数值
    """
    # 处理 None 或空值
    if count_str is None:
        return 0
    
    # 转换为字符串并清理空白
    count_str = str(count_str).strip()
    
    # 如果是空字符串，返回0
    if not count_str:
        return 0
        
    try:
        if '万' in count_str:
            # 去掉 '万'，将字符串转换为浮点数并乘以 10000
            count_str = count_str.replace('万', '')
            return int(float(count_str) * 10000)
        else:
            # 尝试将字符串直接转换为整数
            return int(float(count_str))
    except (ValueError, TypeError):
        # 如果转换失败，返回 0 作为默认值
        return 0

"""处理发布时间函数"""
def parse_time_and_location(note_time_raws):
    """
    解析时间和地理位置信息
    返回格式化的时间 (YYYY-MM-DD) 和位置信息
    """
    # 移除可能存在的"编辑于"前缀
    note_time_raws = note_time_raws.replace('编辑于 ', '')
    
    # 获取当前日期
    current_date = datetime.now()
    
    # 提取位置信息（如果存在）
    location = ''
    location_match = re.search(r'[^\d\s:前天今\-]+$', note_time_raws.strip())
    if location_match:
        location = location_match.group().strip()
        # 从原文本中删除位置信息以便后续处理时间
        note_time_raws = note_time_raws[:note_time_raws.rindex(location)].strip()
    
    # 处理不同的时间格式
    if '天前' in note_time_raws:
        # 处理"X天前"格式
        days = int(re.search(r'(\d+)\s*天前', note_time_raws).group(1))
        result_date = current_date - timedelta(days=days)
    elif '昨天' in note_time_raws:
    # 处理"昨天"格式
        result_date = current_date - timedelta(days=1)
    elif '今天' in note_time_raws:
        # 处理"今天"格式
        result_date = current_date
    elif re.match(r'\d{4}-\d{2}-\d{2}', note_time_raws):
        # 处理"YYYY-MM-DD"格式
        result_date = datetime.strptime(note_time_raws.split()[0], '%Y-%m-%d')
    elif re.match(r'\d{2}-\d{2}', note_time_raws):
        # 处理"MM-DD"格式
        month, day = map(int, note_time_raws.split('-'))
        result_date = datetime(current_date.year, month, day)
        # 如果解析的日期比当前日期晚，说明是去年的日期
        if result_date > current_date:
            result_date = datetime(current_date.year - 1, month, day)
    else:
        raise ValueError(f"Unsupported date format: {note_time_raws}")
    
    # 格式化日期为YYYY-MM-DD
    formatted_date = result_date.strftime('%Y-%m-%d')
    
    return formatted_date, location